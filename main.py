import calendar
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import os


class CalendarGenerator:
    def __init__(self, year):
        self.year = year
        self.events = self.load_predefined_events()

    def load_predefined_events(self):
        # Previous events dictionary remains the same
        events = {
            1: {
                1: ["Revelion", "An nou"],
                24: ["Unirea Principatelor Române"]
            },
            # ... other months ...
        }
        return events

    def create_styled_calendar(self, output_file=None):
        # Generează un nume de fișier unic dacă nu este specificat
        if output_file is None:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f'calendar_{self.year}_{timestamp}.xlsx'

        # Asigură-te că fișierul nu există deja
        if os.path.exists(output_file):
            os.remove(output_file)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for month in range(1, 13):
            ws = wb.create_sheet(title=calendar.month_name[month])

            # Light blue header background
            light_blue = openpyxl.styles.colors.Color(rgb='87CEEB')
            ws.sheet_properties.tabColor = light_blue

            # Header pentru lună
            ws.merge_cells('A1:G2')
            header_cell = ws['A1']
            header_cell.value = calendar.month_name[month].upper()
            header_cell.font = Font(size=16, bold=True)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Antete zile cu fundal ușor albastru
            weekdays = ['SUNDAY', 'MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY']
            for col, day in enumerate(weekdays, start=1):
                cell = ws.cell(row=3, column=col, value=day)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')

            # Generează matricea calendarului
            cal_matrix = calendar.monthcalendar(self.year, month)

            # Populează zilele
            for week_num, week in enumerate(cal_matrix, start=4):
                for day_of_week, day in enumerate(week, start=1):
                    if day == 0:
                        continue

                    cell = ws.cell(row=week_num, column=day_of_week, value=day)
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                    # Border pentru celule
                    thin_border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                    cell.border = thin_border

                    # Adaugă evenimente dacă există
                    if month in self.events and day in self.events[month]:
                        events = self.events[month][day]
                        cell.value = f"{day}\n" + "\n".join(events)
                        cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

            # Lățime coloane
            for col in range(1, 8):
                ws.column_dimensions[get_column_letter(col)].width = 20

            # Adaugă "NOTES" la final
            notes_row = len(cal_matrix) + 5
            ws.merge_cells(f'A{notes_row}:G{notes_row}')
            notes_cell = ws.cell(row=notes_row, column=1, value="NOTES")
            notes_cell.font = Font(bold=True)

            # Adaugă logo vertical 'OCT' pentru fiecare lună
            ws.merge_cells('H1:I10')
            vertical_text_cell = ws['H1']
            vertical_text_cell.value = calendar.month_name[month][:3].upper()
            vertical_text_cell.font = Font(size=48, bold=True)
            vertical_text_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)

        # Salvează și gestionează posibilele erori
        try:
            wb.save(output_file)
            print(f"Calendar pentru {self.year} salvat la: {output_file}")
        except PermissionError:
            print(f"Eroare: Nu se poate salva fișierul {output_file}. Verifică dacă este deschis în altă aplicație.")
        except Exception as e:
            print(f"A apărut o eroare la salvare: {e}")


def main():
    calendar_gen = CalendarGenerator(2025)
    calendar_gen.create_styled_calendar()


if __name__ == "__main__":
    main()