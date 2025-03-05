import calendar
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter
import datetime
import os


class CalendarGenerator:
    def __init__(self, year):
        self.year = year
        self.events = self.load_predefined_events()

    def load_predefined_events(self):
        # Păstrăm evenimentele originale pentru fiecare lună
        events = {
            1: {
                1: ["Revelion", "An nou"],
                24: ["Unirea Principatelor Române"]
            },
            2: {
                14: ["Ziua Îndrăgostiților"]
            },
            3: {
                8: ["Ziua Internațională a Femeii"],
                15: ["Ziua Constituției României"]
            },
            4: {
                15: ["Paște"]
            },
            5: {
                1: ["Ziua Muncii"],
                9: ["Ziua Europei"]
            },
            6: {
                1: ["Ziua Copilului"],
                24: ["Revelație de vară"]
            },
            7: {
                15: ["Zilele culturale de vară"]
            },
            8: {
                15: ["Adormirea Maicii Domnului"]
            },
            9: {
                15: ["Început de toamnă"],
                22: ["Echnocțiul de toamnă"]
            },
            10: {
                1: ["Multe story-uri", "Calea Victoriei"],
                5: ["Seara reel", "Calea Victoriei"],
                7: ["MAP OF THE UNIVERSE"],
                8: ["Real/Tiktok", "telefoane"],
                11: ["Reel/tiktok", "pancarta", "Program orașe"],
                14: ["STORY STAND"],
                15: ["STORY STAND"],
                17: ["Teaser la", "Teaser artist", "filmat pe 16"],
                21: ["STORY STAND", "+Reel artist", "+postare (19/20)"],
                24: ["Story add yours", "editii trecute"],
                26: ["Real/Tiktok", "Smiley"],
                28: ["postare DJ Thomas", "Teaser", "Suna telefonul", "Filmat pe 28"],
                30: ["Giveaway", "tatua"],
                31: ["Postare", "Ce activitate ești?"]
            },
            11: {
                1: ["Început de iarnă"],
                15: ["Ziua Recoltei"]
            },
            12: {
                1: ["Marea Unire"],
                25: ["Crăciun"],
                31: ["Revelion"]
            }
        }
        return events

    def create_styled_calendar(self, output_file=None):
        if output_file is None:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f'calendar_{self.year}_{timestamp}.xlsx'

        if os.path.exists(output_file):
            os.remove(output_file)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Culori definite
        light_blue = '87CEEB'
        event_orange = 'FFA500'

        # Creează o singură foaie pentru întregul an
        ws = wb.create_sheet(title=f"Calendar {self.year}")
        ws.sheet_properties.tabColor = Color(rgb=light_blue)

        # Setează lățimea coloanelor
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Poziție de start pentru primul calendar
        current_row = 1

        for month in range(1, 13):
            # Header pentru lună
            ws.merge_cells(f'A{current_row}:G{current_row + 1}')
            header_cell = ws.cell(row=current_row, column=1, value=calendar.month_name[month].upper())
            header_cell.font = Font(size=16, bold=True)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Antete zile
            weekdays = ['SUNDAY', 'MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY']
            for col, day in enumerate(weekdays, start=1):
                cell = ws.cell(row=current_row + 2, column=col, value=day)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border

            # Generează matricea calendarului
            cal_matrix = calendar.monthcalendar(self.year, month)

            # Populează zilele
            for week_num, week in enumerate(cal_matrix):
                for day_of_week, day in enumerate(week):
                    if day == 0:
                        continue

                    cell = ws.cell(row=current_row + 3 + week_num, column=day_of_week + 1, value=day)
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border

                    # Adaugă evenimente dacă există
                    if month in self.events and day in self.events[month]:
                        events = self.events[month][day]
                        cell.value = f"{day}\n" + "\n".join(events)
                        cell.fill = PatternFill(start_color=event_orange, end_color=event_orange, fill_type='solid')

            # Vertical text pentru lună
            ws.merge_cells(f'H{current_row}:I{current_row + 9}')
            vertical_text_cell = ws.cell(row=current_row, column=8)
            vertical_text_cell.value = calendar.month_name[month][:3].upper()
            vertical_text_cell.font = Font(size=48, bold=True)
            vertical_text_cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                text_rotation=90
            )

            # Adaugă "NOTES"
            notes_row = current_row + len(cal_matrix) + 4
            ws.merge_cells(f'A{notes_row}:G{notes_row}')
            notes_cell = ws.cell(row=notes_row, column=1, value="NOTES")
            notes_cell.font = Font(bold=True)

            # Pregătește poziția pentru luna următoare
            current_row = notes_row + 2

        # Salvează fișierul
        try:
            wb.save(output_file)
            print(f"Calendar pentru {self.year} salvat la: {output_file}")
        except PermissionError:
            print(f"Eroare: Nu se poate salva fișierul {output_file}. Verifică dacă este deschis în altă aplicație.")
        except Exception as e:
            print(f"A apărut o eroare la salvare: {e}")


def main():
    calendar_gen = CalendarGenerator(2025)
    calendar_gen.create_styled_calendar()


if __name__ == "__main__":
    main()